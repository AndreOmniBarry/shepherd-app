#!/usr/bin/env python3
"""
Generates the blank, structure-agnostic church data import template —
handed to a church that wants SHEP.HERD's help importing their existing
records. Works for every structure_type (cell_church, zonal, campus,
single, house_network): tier1/tier2 terminology is deliberately generic
("Group Tier 1" / "Group Tier 2") since the actual label (Fellowship vs.
Zone vs. Campus vs. Department vs. Network) depends on which structure
the church picked at setup — see src/lib/church-config.ts.

Run once whenever the template needs changing:
    python3 scripts/generate_import_template.py
Output: public overwritten at scripts/templates/church_data_import_template.xlsx
(served by /api/admin/import-template, gated behind the consent checkbox
— not linked directly from /public so the consent step can't be skipped).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), 'templates')
OUT_PATH = os.path.join(OUT_DIR, 'church_data_import_template.xlsx')

PURPLE = '534AB7'
PURPLE_DARK = '3C3489'
LIGHT = 'EEEDFE'
WHITE = 'FFFFFF'
TEXT = '0F0A2E'

HEADER_FILL = PatternFill('solid', fgColor=PURPLE)
HEADER_FONT = Font(color=WHITE, bold=True, size=11)
TITLE_FONT = Font(color=PURPLE_DARK, bold=True, size=15)
SUB_FONT = Font(color='4A4272', size=10.5, italic=True)
THIN = Side(style='thin', color='DDDAF5')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def add_sheet(name, columns, notes=None, example_row=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(col) + 4)
    ws.row_dimensions[1].height = 34
    if example_row:
        for i, val in enumerate(example_row, start=1):
            c = ws.cell(row=2, column=i, value=val)
            c.font = Font(italic=True, color='9890C4', size=10)
            c.border = BORDER
    if notes:
        note_row = 4 if example_row else 3
        nc = ws.cell(row=note_row, column=1, value=notes)
        nc.font = SUB_FONT
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(columns))
    ws.freeze_panes = 'A2'
    return ws


# ── Instructions ─────────────────────────────────────────────
ins = wb.create_sheet('Instructions')
ins.sheet_view.showGridLines = False
ins.column_dimensions['A'].width = 100
ins['A1'] = 'SHEP.HERD — Church Data Import Template'
ins['A1'].font = TITLE_FONT
lines = [
    '',
    'This workbook is structure-agnostic — it works whether your church runs Cell Church, Zonal, Campus, '
    'Single Congregation, or House Network. Fill in only the sheets that apply to how your church is organized; '
    'leave the rest blank.',
    '',
    'How to use it:',
    '  1. Fill in "Church Info" first.',
    '  2. Fill "Branches" only if your church has more than one physical location.',
    '  3. Fill "Groups (Tier 1)" and "Groups (Tier 2)" using whatever your church calls them — Fellowship/Cell, '
    'Zone/District, Campus/Fellowship, Department/Unit, or Network/Home Group. Leave Tier 2 blank if your '
    'structure only has one level.',
    '  4. Fill "Departments" for ministry departments (Media, Ushering, Choir, etc.) — separate from the Tier 1/2 '
    'groups above, since a member can belong to both a cell/fellowship AND a department.',
    '  5. Fill "Members" last — one row per person. Use the exact names you typed into the sheets above so '
    'members link up correctly to their group/branch/department.',
    '  6. Fill in the "Consent & Accuracy" sheet — this is required before we can import your data.',
    '',
    'Do not rename the column headers or reorder the sheets — the import script matches by header name and '
    'sheet name.',
    '',
    'When you are done, email the completed file to support@justshephrd.com. Do not upload it anywhere — there '
    'is no upload page for this; a SHEP.HERD team member imports it on your behalf once it arrives and the '
    'Consent & Accuracy sheet is filled in.',
]
for i, line in enumerate(lines, start=2):
    c = ins.cell(row=i, column=1, value=line)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.font = Font(size=11, bold=line.startswith('How to use') or line.startswith('Do not'))
    ins.row_dimensions[i].height = 34 if len(line) > 90 else 18

# ── Church Info ──────────────────────────────────────────────
ci = add_sheet('Church Info', ['Field', 'Your Answer'])
fields = [
    'Church Name', 'Structure Type', 'Country', 'Currency (e.g. NGN, GHS, USD)',
    'Primary Service Day(s)', 'Head Office / Main Branch Address', 'Senior Pastor / Overseer Full Name',
]
for i, f in enumerate(fields, start=2):
    ci.cell(row=i, column=1, value=f).border = BORDER
    ci.cell(row=i, column=2, value='').border = BORDER

dv_structure = DataValidation(type='list', formula1='"Cell Church,Zonal,Campus,Single Congregation,House Network"', allow_blank=True)
ci.add_data_validation(dv_structure)
dv_structure.add(ci.cell(row=3, column=2))

note_row = len(fields) + 3
nc = ci.cell(row=note_row, column=1, value='One row per field. "Structure Type" must be exactly one of: Cell Church, Zonal, Campus, Single Congregation, House Network.')
nc.font = SUB_FONT
ci.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)

# ── Branches ─────────────────────────────────────────────────
add_sheet(
    'Branches',
    ['Branch Name', 'Address', 'Branch Pastor Full Name', 'Branch Pastor Email', 'Branch Pastor Phone'],
    notes='Only needed if your church has more than one physical location/campus. Leave blank for a single-location church.',
    example_row=['Grace Dome (HQ)', '123 Example Street, Lagos', 'Jane Doe', 'jane@example.com', '08012345678'],
)

# ── Groups Tier 1 ────────────────────────────────────────────
add_sheet(
    'Groups (Tier 1)',
    ['Group Name', 'Branch (if multi-location)', 'Leader Full Name', 'Leader Email', 'Leader Phone'],
    notes='Fellowship / Zone / Campus / Department / Network — whichever your church structure uses at the top level.',
    example_row=["Men's Fellowship", 'Grace Dome (HQ)', 'John Smith', 'john@example.com', '08012345678'],
)

# ── Groups Tier 2 ────────────────────────────────────────────
add_sheet(
    'Groups (Tier 2)',
    ['Group Name', 'Parent Tier 1 Group', 'Leader Full Name', 'Leader Email', 'Leader Phone'],
    notes='Cell / District / Fellowship / Unit / Home Group — the level under Tier 1. Leave this whole sheet blank if your structure only has one level (e.g. Single Congregation).',
    example_row=['Cell 3B', "Men's Fellowship", 'Mary Johnson', 'mary@example.com', '08012345678'],
)

# ── Departments ──────────────────────────────────────────────
add_sheet(
    'Departments',
    ['Department Name', 'Head Full Name', 'Head Email', 'Head Phone'],
    notes='Ministry departments — Media, Ushering, Choir, Protocol, etc. Separate from the groups above.',
    example_row=['Media', 'David Lee', 'david@example.com', '08012345678'],
)

# ── Members ──────────────────────────────────────────────────
add_sheet(
    'Members',
    [
        'Full Name', 'Gender', 'Phone', 'Email (optional)', 'Address', 'Date of Birth',
        'Marital Status', 'Occupation', 'Join Date', 'Branch', 'Tier 1 Group', 'Tier 2 Group',
        'Department(s) — comma-separated', 'Membership Status',
    ],
    notes='One row per person. Use the exact same names as typed in the sheets above so each person links to the right branch/group/department.',
    example_row=[
        'Grace Adeyemi', 'Female', '08012345678', 'grace@example.com', '45 Example Close, Lagos',
        '1990-04-12', 'Married', 'Teacher', '2019-06-01', 'Grace Dome (HQ)', "Women's Fellowship",
        'Cell 3B', 'Media, Choir', 'Active',
    ],
)


# ── Consent & Accuracy ───────────────────────────────────────
# This is the actual liability handoff point — the church is the one
# attesting the data is accurate, not SHEP.HERD. The in-app download
# checkbox only logs that a SHEP.HERD admin handed the blank template out;
# THIS is what matters, since it travels with the filled file back over
# email and is what a church rep actually fills in and signs off on.
cs = wb.create_sheet('Consent & Accuracy')
cs.sheet_view.showGridLines = False
cs.column_dimensions['A'].width = 100
cs['A1'] = 'Consent & Data Accuracy'
cs['A1'].font = TITLE_FONT
consent_lines = [
    '',
    'By completing and returning this file to SHEP.HERD, the person named below confirms, on behalf of the '
    'church, that:',
    '  • The information in this file has been reviewed for accuracy by someone with authority to do so.',
    '  • SHEP.HERD is not responsible for errors, omissions, or outdated information contained in the data '
    'provided in this file — the church remains responsible for the accuracy of its own records.',
    '  • SHEP.HERD will import this data as provided, and the church can correct any entry afterward from '
    'within the app.',
    '',
    'Please fill in and return with the completed sheets:',
]
for i, line in enumerate(consent_lines, start=2):
    c = cs.cell(row=i, column=1, value=line)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    cs.row_dimensions[i].height = 34 if len(line) > 90 else 18

sign_start = 2 + len(consent_lines) + 1
sign_fields = ['Full Name', 'Title / Role at Church', 'Church Name', 'Date']
cs.column_dimensions['B'].width = 40
for i, f in enumerate(sign_fields, start=sign_start):
    lbl = cs.cell(row=i, column=1, value=f)
    lbl.font = Font(bold=True, size=11, color=PURPLE_DARK)
    val = cs.cell(row=i, column=2, value='')
    val.border = Border(bottom=Side(style='thin', color=PURPLE))

os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT_PATH)
print(f'Wrote {OUT_PATH}')
