#!/usr/bin/env python3
"""
One-time generator: reads the Grace Dome Excel export and produces
scripts/12_import_grace_dome_data.sql (fellowships/cells/departments/members)
and scripts/leaders_for_accounts.json (input for create-leader-accounts.mjs).

Run once: python3 scripts/generate_import.py
"""
import openpyxl, json, re, sys

SRC = '/root/.claude/uploads/a8365774-cb99-5ffe-8a87-f26cca2dc84f/53a487ca-TCH_Grace_Dome_DB2_1.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def esc(s):
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''").strip() + "'"

def norm_phone(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', '-', 'None'):
        return None
    s = re.sub(r'\.0$', '', s)
    s = re.sub(r'\D', '', s)
    if not s:
        return None
    if len(s) == 10 and not s.startswith('0'):
        s = '0' + s
    if len(s) == 13 and s.startswith('234'):
        s = '0' + s[3:]
    return s

FELLOWSHIP_MAP = {
    'youth': 'Youth Fellowship', 'youth fellowship': 'Youth Fellowship',
    'power house (youth)': 'Youth Fellowship', 'youth/teen': 'Youth Fellowship',
    'men': "Men's Fellowship", 'man': "Men's Fellowship", 'men fellowship': "Men's Fellowship",
    "men's fellowship": "Men's Fellowship",
    'women': "Women's Fellowship", 'woman fellowship': "Women's Fellowship",
    "women's fellowship": "Women's Fellowship",
    'teen': "Teenager's Fellowship", 'teenage': "Teenager's Fellowship",
    'teenager': "Teenager's Fellowship", 'teenager fellowship': "Teenager's Fellowship",
    'teenagers fellowship': "Teenager's Fellowship", 'teens fellowship': "Teenager's Fellowship",
    'teenage fellowship': "Teenager's Fellowship",
}
def norm_fellowship(v):
    if not v:
        return None
    key = str(v).strip().lower()
    return FELLOWSHIP_MAP.get(key)  # None (unmapped: 'c t h', 'creative art', 'n/a', etc.) -> unassigned

sql = []
sql.append("-- ============================================================")
sql.append("-- Grace Dome data import — generated from TCH_Grace_Dome_DB2_1.xlsx")
sql.append("-- Run once in the Supabase SQL editor, top to bottom.")
sql.append("-- Idempotent-ish: uses ON CONFLICT DO NOTHING / WHERE NOT EXISTS guards")
sql.append("-- so it is safe to re-run if it fails partway through.")
sql.append("-- ============================================================\n")

sql.append("-- Schema additions needed for the Create Member feature (address/occupation)")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS address TEXT;")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS occupation TEXT;")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS marital_status TEXT;")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS emergency_contact TEXT;")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS next_of_kin TEXT;")
sql.append("ALTER TABLE members ADD COLUMN IF NOT EXISTS baptism_status TEXT;\n")

# ── Fellowships ──────────────────────────────────────────────
fh_ws = wb['Fellowship Heads']
fellowships = []  # (name, head_title_col_value)
for row in fh_ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    dept_col = str(row[3] or '').replace(' Head', '').strip()
    fellowships.append(dept_col)
# Ensure canonical 4 always present even if sheet text differs
for canon in ["Youth Fellowship", "Men's Fellowship", "Women's Fellowship", "Teenager's Fellowship"]:
    if canon not in fellowships:
        fellowships.append(canon)

sql.append("-- ── Fellowships ──────────────────────────────────────────────")
for f in fellowships:
    sql.append(f"INSERT INTO fellowships (name) SELECT {esc(f)} WHERE NOT EXISTS (SELECT 1 FROM fellowships WHERE name = {esc(f)});")
sql.append("")

# ── Cells ────────────────────────────────────────────────────
cl_ws = wb['Cell Leaders']
cells = []  # (cell_name, fellowship_name)
for row in cl_ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    cell_name = str(row[3]).strip()
    fel = norm_fellowship(row[4]) or "Youth Fellowship"
    cells.append((cell_name, fel))

sql.append("-- ── Cells ────────────────────────────────────────────────────")
for cell_name, fel in cells:
    sql.append(
        f"INSERT INTO cells (name, fellowship_id, is_active) "
        f"SELECT {esc(cell_name)}, f.id, true FROM fellowships f WHERE f.name = {esc(fel)} "
        f"AND NOT EXISTS (SELECT 1 FROM cells c WHERE c.name = {esc(cell_name)} AND c.fellowship_id = f.id);"
    )
sql.append("")

# ── Departments ──────────────────────────────────────────────
dh_ws = wb['Dept Heads']
departments = []
for row in dh_ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    dept_name = str(row[3] or '').replace(' Head', '').strip()
    departments.append(dept_name)
# Choir appears in Dept Members with no head row — add explicitly (led informally under Music)
if 'Choir' not in departments:
    departments.append('Choir')

sql.append("-- ── Departments ──────────────────────────────────────────────")
for d in departments:
    sql.append(f"INSERT INTO departments (name) SELECT {esc(d)} WHERE NOT EXISTS (SELECT 1 FROM departments WHERE name = {esc(d)});")
sql.append("")

# ── Members (Members Register) ──────────────────────────────
mr_ws = wb['Members Register']
seen = set()
member_rows = []
for row in mr_ws.iter_rows(min_row=2, values_only=True):
    name = row[1]
    if not name:
        continue
    name = str(name).strip()
    phone = norm_phone(row[3])
    key = (name.lower(), phone)
    if key in seen:
        continue
    seen.add(key)
    address = row[2]
    fel = norm_fellowship(row[4])
    member_rows.append((name, address, phone, fel))

sql.append(f"-- ── Members Register ({len(member_rows)} unique members) ──────────────")
for name, address, phone, fel in member_rows:
    fel_sub = f"(SELECT id FROM fellowships WHERE name = {esc(fel)} LIMIT 1)" if fel else "NULL"
    sql.append(
        f"INSERT INTO members (full_name, address, phone, fellowship_id, membership_status, join_date) "
        f"SELECT {esc(name)}, {esc(address)}, {esc(phone)}, {fel_sub}, 'active', CURRENT_DATE "
        f"WHERE NOT EXISTS (SELECT 1 FROM members m WHERE m.full_name = {esc(name)} "
        f"AND ({'m.phone = ' + esc(phone) if phone else 'm.phone IS NULL'}));"
    )
sql.append("")

# ── Department members (from Dept Members sheet + dept heads) ──
dm_ws = wb['Dept Members']
DEPT_NORM = {
    'media department': 'Media', 'media hod': 'Media',
    'lighting department': 'Lighting', 'lighting hod': 'Lighting',
    'protocol department': 'Protocol', 'protocol hod': 'Protocol',
    'choir': 'Choir',
    'prayer': 'Prayer', 'prayer ': 'Prayer',
    'creative art': 'Creative Art',
}
dept_member_rows = []  # (name, phone, dept_name, role_in_dept)
for row in dm_ws.iter_rows(min_row=2, values_only=True):
    name = row[1]
    if not name:
        continue
    name = str(name).strip()
    role = str(row[2] or 'member').strip()
    dept_raw = str(row[3] or '').strip().lower()
    dept_name = DEPT_NORM.get(dept_raw, str(row[3] or '').strip())
    phone = norm_phone(row[4])
    dept_member_rows.append((name, phone, dept_name, role))

sql.append(f"-- ── Department members ({len(dept_member_rows)} rows, matched/created against members) ──")
for name, phone, dept, role in dept_member_rows:
    phone_clause = f"m.phone = {esc(phone)}" if phone else "m.phone IS NULL"
    # Ensure a members row exists for this person (dept-only members not in Members Register)
    sql.append(
        f"INSERT INTO members (full_name, phone, membership_status, join_date) "
        f"SELECT {esc(name)}, {esc(phone)}, 'active', CURRENT_DATE "
        f"WHERE NOT EXISTS (SELECT 1 FROM members m WHERE m.full_name = {esc(name)} AND {phone_clause});"
    )
    sql.append(
        f"INSERT INTO department_members (department_id, member_id, role) "
        f"SELECT d.id, m.id, {esc(role)} FROM departments d, members m "
        f"WHERE d.name = {esc(dept)} AND m.full_name = {esc(name)} AND {phone_clause} "
        f"AND NOT EXISTS (SELECT 1 FROM department_members dm2 WHERE dm2.department_id = d.id AND dm2.member_id = m.id);"
    )
sql.append("")

with open('/home/user/shepherd-app/scripts/12_import_grace_dome_data.sql', 'w') as f:
    f.write('\n'.join(sql))

print(f"Wrote scripts/12_import_grace_dome_data.sql — {len(sql)} statements")
print(f"Fellowships: {len(fellowships)}, Cells: {len(cells)}, Departments: {len(departments)}, Members: {len(member_rows)}, Dept-member links: {len(dept_member_rows)}")

# ── Leadership accounts JSON (for the Node account-creation script) ──
leaders = []

def add_leader(sheet_name, role, ref_col_idx, ref_kind):
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        name = str(row[1]).strip()
        email = row[4] if len(row) > 4 else None
        if not email or '@' not in str(email):
            slug = re.sub(r'[^a-z.]', '', name.lower().replace(' ', '.'))
            email = f"{slug}@shepherd.app"
        phone = norm_phone(row[5]) if len(row) > 5 else None
        ref_name = str(row[ref_col_idx]).strip() if ref_col_idx is not None and row[ref_col_idx] else None
        leaders.append({
            'name': name, 'email': str(email).strip().lower(), 'phone': phone,
            'role': role, 'ref_kind': ref_kind, 'ref_name': ref_name,
        })

add_leader('G.O', 'overseer', None, None)
add_leader('Pastor & Admins', None, None, None)  # role resolved per-row below (mixed roles in this sheet)
add_leader('Tech', 'lead_tech', None, None)
add_leader('Fellowship Heads', 'fellowship_head', 3, 'fellowship')
add_leader('Dept Heads', 'department_head', 3, 'department')
add_leader('Cell Leaders', 'cell_leader', 3, 'cell')

# Fix Pastor & Admins roles individually (this sheet mixes accounts/partnership/pa roles)
PA_ROLE_MAP = {
    'church admin - pa to bishop': 'pa',
    'church accountant': 'accounts',
    'church partner admin': 'partnership',
    'assistant pastor - grace dome': 'pa',
    'residing pastor - victory tarbenacle': 'pa',
}
pa_ws = wb['Pastor & Admins']
pa_idx = 0
for i, l in enumerate(leaders):
    if l['role'] is None:
        row = list(pa_ws.iter_rows(min_row=2 + pa_idx, max_row=2 + pa_idx, values_only=True))[0]
        branch = str(row[3] or '').strip().lower()
        l['role'] = PA_ROLE_MAP.get(branch, 'pa')
        pa_idx += 1

# Fold in for cells/fellowships whose canonical name needs the Fellowship suffix normalization
for l in leaders:
    if l['ref_kind'] == 'fellowship' and l['ref_name']:
        l['ref_name'] = norm_fellowship(l['ref_name']) or l['ref_name']
    if l['ref_kind'] == 'cell' and l['ref_name']:
        pass  # cell names already canonical from Cell Leaders sheet
    if l['ref_kind'] == 'department' and l['ref_name']:
        l['ref_name'] = DEPT_NORM.get(l['ref_name'].strip().lower(), l['ref_name'])

with open('/home/user/shepherd-app/scripts/leaders_for_accounts.json', 'w') as f:
    json.dump(leaders, f, indent=2)

print(f"Wrote scripts/leaders_for_accounts.json — {len(leaders)} leadership accounts")
